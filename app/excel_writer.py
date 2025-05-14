"""Code for logging results.

This module provides functions to:
  * Initialize or load the Excel report workbook.
  * Append timestamped test results (MAC, status) to the workbook.

Public API:
    initialize_workbook
    append_result
"""
# report_logger.py

import os
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException, WorkbookAlreadySaved
from openpyxl.styles import PatternFill

from log import logger  # assumes your log.py exposes 'logger'

# Constants
REPORT_FILE = "wifi_test_results.xlsx"
GREEN_FILL = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
RED_FILL   = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

def initialize_workbook() -> bool:
    """
    Ensure REPORT_FILE exists with headers.
    Returns True if workbook is ready, False on failure.
    """
    if not os.path.exists(REPORT_FILE):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "TestResults"
            ws.append(["Timestamp", "MAC", "Status"])
            wb.save(REPORT_FILE)
            logger.info("Created new workbook '%s' with headers.", REPORT_FILE)
            return True
        except (OSError, IOError) as e:
            logger.error("Filesystem error creating '%s'.", REPORT_FILE, exc_info=True)
            return False
        except Exception as e:
            logger.error("Unexpected error initializing workbook.", exc_info=True)
            return False
    else:
        logger.info("Workbook '%s' already exists.", REPORT_FILE)
        return True

def append_result(mac: str, status: str) -> bool:
    """
    Append a timestamped test result to REPORT_FILE.
    Returns True on success, False on any error.
    """
    # Ensure workbook structure first
    if not initialize_workbook():
        return False

    try:
        wb = load_workbook(REPORT_FILE)
        ws = wb.active
    except InvalidFileException as e:
        logger.error("Invalid Excel file '%s'.", REPORT_FILE, exc_info=True)
        return False
    except WorkbookAlreadySaved as e:
        logger.error("Attempted to re-save a write-only workbook '%s'.", REPORT_FILE, exc_info=True)
        return False
    except (OSError, IOError) as e:
        logger.error("Filesystem error loading '%s'.", REPORT_FILE, exc_info=True)
        return False
    except Exception as e:
        logger.error("Unexpected error loading workbook.", exc_info=True)
        return False

    # Append the data row
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        ws.append([timestamp, mac, status])
        # Apply colored fill to the latest status cell
        cell = ws.cell(row=ws.max_row, column=3)
        cell.fill = GREEN_FILL if status.lower() == "pass" else RED_FILL
        wb.save(REPORT_FILE)
        logger.info("Appended result for MAC=%s, status=%s at %s.", mac, status, timestamp)
        return True
    except (OSError, IOError) as e:
        logger.error("Filesystem error saving '%s'.", REPORT_FILE, exc_info=True)
        return False
    except Exception as e:
        logger.error("Unexpected error appending result.", exc_info=True)
        return False
