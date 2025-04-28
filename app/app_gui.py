import os
import tkinter as tk
from tkinter import Menu, messagebox
import tkinter.font as tkFont
import pyudev
import subprocess
import netifaces
from openpyxl import Workbook, load_workbook
from datetime import datetime
from tkinter import ttk
import time # For testing purposes, simulate connection time

REPORT_FILE = "wifi_test_results.xlsx"

class HardwareTestApp:
    def __init__(self, root):
        self.root = root
        self.root.title("WiFi Test Application")
        self.root.minsize(800, 600)

        # Configurable parameters
        self.ssid = "ssid"
        self.password = "psswd"

        # Test data
        self.interfaces = []      # list of iface names
        self.test_results = {}    # iface -> "PASS"/"FAIL"
        self.buttons = {}         # iface -> Button widget
        self.test_buttons = {}    # iface -> Test Button widget

        self.create_menu()
        self.create_widgets()
        self.refresh_interfaces()

    def create_menu(self):
        """Creates a File menu with options to open/save results, configure test parameters, and access help."""
        menu_bar = Menu(self.root)
        
        # File menu
        file_menu = Menu(menu_bar, tearoff=0)
        file_menu.add_command(label="Save Results", command=self.save_results)
        file_menu.add_separator()
        file_menu.add_command(label="Reset", command=self.reset_tests)
        file_menu.add_command(label="Exit", command=self.root.quit)
        menu_bar.add_cascade(label="File", menu=file_menu)

        # Configure menu
        menu_bar.add_command(label="Configure", command=self.configure_parameters)

        # Help Menu
        help_menu = Menu(menu_bar, tearoff=0)
        help_menu.add_command(label="Help", command=self.show_help)
        # help_menu.add_command(label="About", command=self.show_about)
        help_menu.add_separator()
        # help_menu.add_command(label="Contact Support", command=self.contact_support)
        menu_bar.add_cascade(label="Help", menu=help_menu)
        
        self.root.config(menu=menu_bar)

    def create_widgets(self):
        main = tk.Frame(self.root)
        main.pack(fill=tk.BOTH, expand=True)
        heading_font = tkFont.Font(family="Helvetica", size=11, weight="normal")

        # Left pane for interface buttons
        left = tk.Frame(main, bd=2, relief=tk.SUNKEN)
        left.pack(side=tk.LEFT, fill=tk.Y, padx=5, pady=5)
        self.left_frame = left

        # -- Heading + LED indicator canvas ------------------
        header_frame = tk.Frame(self.left_frame)
        header_frame.pack(fill=tk.X, pady=(0,10), padx=0)

        lbl = tk.Label(
            header_frame,
            text="Detected Devices",
            pady=5,
            font=heading_font,
            bg="#e0e0e0", bd=1, relief=tk.SOLID
        )
        lbl.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # LED: tiny canvas for colored dot
        self.led = tk.Canvas(
            header_frame, width=25, height=25, highlightthickness=0, 
            bg="#e0e0e0", bd=1, relief=tk.SOLID)
        # draw circle centered in that 16×16
        self.led_dot = self.led.create_oval(6,6,20,20, fill="red")   # default red
        self.led.pack(side=tk.RIGHT)

        # Right pane for logs
        right = tk.Frame(main, bd=2, relief=tk.SUNKEN)
        right.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        tk.Label(right, text="Log").pack(anchor="w")
        self.log = tk.Text(right, wrap=tk.WORD, height=15, state=tk.DISABLED)
        self.log.pack(fill=tk.BOTH, expand=True)

        # Progress bar (below the Run Tests button)
        self.progress_var = tk.DoubleVar()
        self.progress = ttk.Progressbar(
            self.root,
            variable=self.progress_var,
            maximum=1,            # will reset later to number of interfaces
            mode='determinate',   # shows exact progress
        )
        self.progress.pack(fill=tk.X, padx=5, pady=5)

        # Run Tests button
        self.run_button = tk.Button(self.root, text="Test All", command=self.run_all_tests)
        self.run_button.pack(pady=5)

    def log_message(self, msg):
        self.log.config(state=tk.NORMAL)
        self.log.insert(tk.END, msg + "\n")
        self.log.see(tk.END)
        self.log.config(state=tk.DISABLED)
        # Force GUI to update
        self.root.update_idletasks()
    
    def clear_log(self):
        self.log.configure(state=tk.NORMAL)
        self.log.delete("1.0", tk.END)
        self.log.configure(state=tk.DISABLED)
        self.progress_var.set(0)    # reset progress bar

    def reset_tests(self):
        """Reset all tests for the next device."""
        self.clear_log()

    def configure_parameters(self):
        """Popup a window to allow user to edit MAC address, serial port, and model number simultaneously."""
        
        # Create a new Toplevel window (acts as a modal dialog)
        window = tk.Toplevel(self.root)
        window.title("Configure Test Parameters")
        
        # --- SSID field ---
        tk.Label(window, text="WiFi SSID:").grid(row=1, column=0, sticky="e", padx=5, pady=5)
        ssid_entry = tk.Entry(window, width=30)
        ssid_entry.insert(0, self.ssid)
        ssid_entry.grid(row=1, column=1, padx=5, pady=5)
        
        # --- Password field ---
        tk.Label(window, text="Password:").grid(row=2, column=0, sticky="e", padx=5, pady=5)
        password_entry = tk.Entry(window, width=30)
        password_entry.insert(0, self.password)
        password_entry.grid(row=2, column=1, padx=5, pady=5)
        
        # --- OK Button to save changes ---
        def on_ok():
            self.ssid = ssid_entry.get()
            self.password = password_entry.get()
            # Optionally, you can show a message box or log the changes
            # For example:
            # messagebox.showinfo("Parameters Set",
            #                     f"MAC Address: {self.mac_addr}\n"
            #                     f"IP Address: {self.server_ip}\n"
            #                     f"Serial Port: {self.serial_port}\n"
            #                     f"Model Number: {self.model_number}")
            window.destroy()

        tk.Button(window, text="OK", command=on_ok).grid(row=3, column=0, columnspan=2, pady=10)
        
        # Update window to ensure its size is computed.
        window.update_idletasks()

        # Calculate position: top center relative to self.root
        root_x = self.root.winfo_x()
        root_y = self.root.winfo_y()
        root_width = self.root.winfo_width()
        win_width = window.winfo_reqwidth()
        # Place the popup centered horizontally relative to the main window,
        # and at the top (or a few pixels offset from the top).
        x = root_x + (root_width // 2) - (win_width // 2)
        y = root_y + 10  # 10 pixels below the top edge of the main window
        window.geometry(f"+{x}+{y}")

        # Make the window modal
        window.grab_set()
    
    def show_help(self):
        """Display help information."""
        messagebox.showinfo(
            "Help",  # Title in title case, no end punctuation
            "This application tests USB-connected MT7601U Wi-Fi adapters by connecting each to a configured network and verifying IP assignment.\n\n"  # Sentence style, explains purpose clearly :contentReference[oaicite:1]{index=1}
            "• File → Open Results: load a previously saved Excel log.\n"             # Imperative verb, concise
            "• File → Save Results: write current test outcomes to the report file.\n"  # Imperative verb, consistent style
            "• Configure: set Wi-Fi SSID, password, and serial-port parameters.\n"       # Noun phrase, action is obvious
            "• Run Tests: detect all connected adapters, attempt connections, and check for IPs.\n"  # Descriptive, accurate
            "• Progress bar: shows live count of completed tests; updates in real time.\n"         # Body text supports title, secondary detail
            "• Logs: view step-by-step status in the text area for troubleshooting.\n"               # Avoids technical jargon, tells user where to look
            "• Adapter Buttons: click an individual interface to retest a single dongle.\n"           # Consistent with button labels, clear action
            "• Reconnect: re-scan USB and serial/Wi-Fi interfaces on demand.\n"                      # Explicit label, explains outcome
            "• Color codes: green=PASS, red=FAIL; results saved automatically to Excel.\n"            # Respects user effort, provides constructive feedback 
            "• Close this dialog by clicking OK or the × in the corner.\n"                           # Provides visible close option
        )

    def refresh_interfaces(self):
        self.clear_log()
        self.log_message("Refreshing interfaces…")

        # 2) enumerate only 'net' devices that already carry the VID/PID properties
        ctx = pyudev.Context()
        # list_devices(**kwargs) forwards to match_subsystem() & match_property() :contentReference[oaicite:4]{index=4}
        matches = list(ctx.list_devices(
            subsystem="net",
            ID_VENDOR_ID="148f",
            ID_MODEL_ID="7601"
        ))

        # 3) collect interface names
        self.interfaces = [dev.sys_name for dev in matches]
        # self.interfaces = ['wlan0', 'wlan1']  # For testing purposes, hardcoded to wlan0 and wlan1 

        # 4) debug-log exactly what we found
        if not self.interfaces:
            self.log_message("No MT7601U interfaces found.")
            self.led.itemconfig(self.led_dot, fill="red")  # red
        else:
            self.log_message(f"Found {len(self.interfaces)} WiFi devices: {self.interfaces}")
            self.led.itemconfig(self.led_dot, fill="green")  # green

        # Clear old buttons
        for widget in self.left_frame.winfo_children():
            if isinstance(widget, tk.Button):
                widget.destroy()

        # Create one button per interface
        for iface in self.interfaces:
            btn = tk.Button(
                self.left_frame,
                text=iface,
                width=15,
                command=lambda i=iface: self.test_interface(i)  # bind current iface
            )
            btn.pack(pady=2)
            self.test_buttons[iface] = btn

    def test_interface(self, iface):
        self.log_message(f"Testing {iface}...")
        # Connect
        ok = self.connect_wifi(iface, self.ssid, self.password)
        self.log_message(f"{iface} WiFi connection {'succeeded' if ok else 'failed'}")
        # Verify IP
        passed = ok and self.check_ip(iface)
        # time.sleep(2)  # Simulate connection time 
        # passed = True
        status = "PASS" if passed else "FAIL"
        self.test_results[iface] = status
        self.log_message(f"{iface}: {status}")
        append_result(self.get_mac(iface), status)
        # Update button color
        #color = "green" if passed else "red"
        #self.buttons[iface].config(bg=color)

    def run_all_tests(self):
        self.refresh_interfaces()
        if not self.interfaces:
            messagebox.showwarning("No Adapters", "No MT7601U adapters found.")
            return
        # reset progress bar
        total = len(self.interfaces)
        self.progress.config(maximum=total)
        self.progress_var.set(0)

        for iface in self.interfaces:
            self.test_interface(iface)
            # time.sleep(1)  # Simulate time taken for each test
            # advance progress bar immediately
            self.progress_var.set(self.progress_var.get() + 1)
            self.root.update_idletasks()        # ensure bar moves in real time

    def save_results(self):
        messagebox.showinfo("Saved", f"Results saved to {REPORT_FILE}")

    @staticmethod
    def connect_wifi(iface, ssid, password):
        try:
            subprocess.run(["nmcli", "radio", "wifi", "on"], check=True)
            subprocess.run(["nmcli", "device", "wifi", "rescan"], check=True)
            subprocess.run(
                ["nmcli", "device", "wifi", "connect", ssid, "password", password, "ifname", iface],
                check=True, timeout=20
            )
            return True
        except subprocess.SubprocessError:
            return False

    @staticmethod
    def check_ip(iface):
        addrs = netifaces.ifaddresses(iface)
        return netifaces.AF_INET in addrs and bool(addrs[netifaces.AF_INET])

    @staticmethod
    def get_mac(iface):
        addrs = netifaces.ifaddresses(iface)
        link = addrs.get(netifaces.AF_LINK)
        return link[0]['addr'] if link else "Unknown"

def initialize_workbook():
    if not os.path.exists(REPORT_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "TestResults"
        ws.append(["Timestamp", "MAC Address", "Status"])
        wb.save(REPORT_FILE)

def append_result(mac, status):
    initialize_workbook()
    wb = load_workbook(REPORT_FILE)
    ws = wb.active
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([ts, mac, status])
    wb.save(REPORT_FILE)

if __name__ == "__main__":
    initialize_workbook()
    root = tk.Tk()
    app = HardwareTestApp(root)
    root.mainloop()
